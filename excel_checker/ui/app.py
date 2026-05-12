import os
import re
import sys
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import threading

from openpyxl import load_workbook # pyright: ignore[reportMissingModuleSource]

from excel_checker.config import (
    HEADER_ROW,
    APP_MIN_WIDTH,
    APP_MIN_HEIGHT,
    EMAIL_REGEX_DEFAULT,
)
from excel_checker.services.column_detector import (
    build_detection_mode_label,
    find_target_columns,
    get_all_non_empty_columns,
    hex_to_rgb,
)
from excel_checker.services.explanation_parser import (
    parse_explanation_workbook,
    classify_column_with_explanation,
)
from excel_checker.services.report_generator import generate_report
from excel_checker.utils.text_utils import clean_header_name, parse_header_metadata
from excel_checker.utils.validation import is_email_header

class ExcelMandatoryCheckerApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Contrôle des colonnes obligatoires Excel - V1.0.0")
        self.root.geometry("1450x1020")
        self.root.minsize(APP_MIN_WIDTH, APP_MIN_HEIGHT)

        self.file_path = None
        self.explanation_file_path = None
        self.sheet_names = []

        self.detected_columns_by_sheet = {}
        self.column_vars_by_sheet = {}
        self.column_categories_by_sheet = {}

        self.current_sheet_for_columns = None

        self.explanation_data = {
            "mandatory": set(),
            "recommended": set(),
            "by_group": {}
        }

        self._build_ui()
        self._bind_mousewheel()

    def resource_path(self, relative_path):
        if getattr(sys, "frozen", False):
            base_path = sys._MEIPASS
        else:
            base_path = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", ".."))
        full_path = os.path.join(base_path, relative_path)
        print("ICON PATH =", full_path)
        print("EXISTS =", os.path.exists(full_path))
        return full_path

    def _bind_mousewheel(self):
        self.root.bind_all("<MouseWheel>", self._on_mousewheel)      # Windows
        self.root.bind_all("<Button-4>", self._on_mousewheel_linux) # Linux
        self.root.bind_all("<Button-5>", self._on_mousewheel_linux) # Linux

    def _on_mousewheel(self, event):
        self.main_canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")

    def _on_mousewheel_linux(self, event):
        if event.num == 4:
            self.main_canvas.yview_scroll(-1, "units")
        elif event.num == 5:
            self.main_canvas.yview_scroll(1, "units")

    def _build_ui(self):
        self.root.rowconfigure(0, weight=1)
        self.root.columnconfigure(0, weight=1)
        self.root.iconbitmap(self.resource_path("app2.ico"))

        # Conteneur principal scrollable
        outer_frame = ttk.Frame(self.root)
        outer_frame.grid(row=0, column=0, sticky="nsew")
        outer_frame.rowconfigure(0, weight=1)
        outer_frame.columnconfigure(0, weight=1)

        bg_color = self.root.cget("background")

        self.main_canvas = tk.Canvas(
            outer_frame,
            highlightthickness=0,
            bd=0,
            bg=bg_color
        )
        self.main_canvas.grid(row=0, column=0, sticky="nsew")

        self.main_scrollbar = ttk.Scrollbar(
            outer_frame,
            orient="vertical",
            command=self.main_canvas.yview
        )
        self.main_scrollbar.grid(row=0, column=1, sticky="ns")

        self.main_canvas.configure(yscrollcommand=self.main_scrollbar.set)

        self.content_frame = ttk.Frame(self.main_canvas)
        self.main_window = self.main_canvas.create_window(
            (0, 0),
            window=self.content_frame,
            anchor="nw"
        )

        self.content_frame.bind("<Configure>", self._on_main_frame_configure)
        self.main_canvas.bind("<Configure>", self._on_main_canvas_configure)

        # Ancien "main"
        main = ttk.Frame(self.content_frame, padding=12)
        main.grid(row=0, column=0, sticky="nsew")

        self.content_frame.columnconfigure(0, weight=1)
        self.content_frame.rowconfigure(0, weight=1)

        main.rowconfigure(4, weight=1)
        main.columnconfigure(0, weight=1)

        file_frame = ttk.LabelFrame(main, text="1) Fichier Excel à contrôler", padding=10)
        file_frame.grid(row=0, column=0, sticky="ew", pady=(0, 10))
        file_frame.columnconfigure(0, weight=1)

        self.file_var = tk.StringVar()

        ttk.Entry(file_frame, textvariable=self.file_var, state="readonly").grid(
            row=0, column=0, sticky="ew", padx=(0, 8)
        )

        ttk.Button(
            file_frame,
            text="Choisir un fichier Excel",
            command=self.choose_file
        ).grid(row=0, column=1, sticky="e")

        explanation_frame = ttk.LabelFrame(main, text="2) Fichier d'explication client", padding=10)
        explanation_frame.grid(row=1, column=0, sticky="ew", pady=(0, 10))
        explanation_frame.columnconfigure(0, weight=1)

        self.explanation_file_var = tk.StringVar()

        ttk.Entry(
            explanation_frame,
            textvariable=self.explanation_file_var,
            state="readonly"
        ).grid(row=0, column=0, sticky="ew", padx=(0, 8))

        ttk.Button(
            explanation_frame,
            text="Charger le fichier d'explication",
            command=self.choose_explanation_file
        ).grid(row=0, column=1, sticky="e")

        self.use_explanation_file_var = tk.BooleanVar(value=True)
        ttk.Checkbutton(
            explanation_frame,
            text="Utiliser le fichier d'explication pour classer les colonnes en Obligatoire / Recommandée",
            variable=self.use_explanation_file_var,
            command=self.refresh_detected_columns
        ).grid(row=1, column=0, columnspan=2, sticky="w", pady=(8, 0))

        self.explanation_info_var = tk.StringVar(value="Aucun fichier d'explication chargé.")
        ttk.Label(explanation_frame, textvariable=self.explanation_info_var).grid(
            row=2, column=0, columnspan=2, sticky="w", pady=(6, 0)
        )

        self.param_frame = ttk.LabelFrame(main, text="3) Paramètres de détection", padding=10)
        self.param_frame.grid(row=2, column=0, sticky="ew", pady=(0, 10))

        for i in range(12):
            self.param_frame.columnconfigure(i, weight=0)
        self.param_frame.columnconfigure(10, weight=1)

        ttk.Label(self.param_frame, text="Ligne des entêtes :").grid(row=0, column=0, sticky="w")
        self.header_row_var = tk.StringVar(value=str(HEADER_ROW))
        ttk.Entry(self.param_frame, textvariable=self.header_row_var, width=8).grid(
            row=0, column=1, sticky="w", padx=(8, 14)
        )

        ttk.Label(self.param_frame, text="Détection classique :").grid(row=0, column=2, sticky="w", padx=(0, 8))

        self.detect_by_color_var = tk.BooleanVar(value=True)
        self.detect_by_star_var = tk.BooleanVar(value=False)

        ttk.Checkbutton(
            self.param_frame,
            text="Par couleur",
            variable=self.detect_by_color_var
        ).grid(row=0, column=3, sticky="w", padx=(0, 8))

        ttk.Checkbutton(
            self.param_frame,
            text="Par *",
            variable=self.detect_by_star_var
        ).grid(row=0, column=4, sticky="w", padx=(0, 16))

        ttk.Button(
            self.param_frame,
            text="Détecter / recharger les colonnes",
            command=self.refresh_detected_columns
        ).grid(row=0, column=11, sticky="e")

        ttk.Label(self.param_frame, text="Couleur hex :").grid(row=1, column=0, sticky="w", pady=(8, 0))

        self.custom_color_var = tk.StringVar(value="FFFF00")
        ttk.Entry(self.param_frame, textvariable=self.custom_color_var, width=10).grid(
            row=1, column=1, sticky="w", padx=(8, 14), pady=(8, 0)
        )

        self.use_custom_color_var = tk.BooleanVar(value=False)
        ttk.Checkbutton(
            self.param_frame,
            text="Utiliser cette couleur personnalisée",
            variable=self.use_custom_color_var
        ).grid(row=1, column=2, columnspan=3, sticky="w", pady=(8, 0))

        ttk.Label(
            self.param_frame,
            text=(
                "Si le fichier d'explication est activé, il sert à classer les colonnes trouvées. "
                "Sinon, le comportement V6/V5 reste disponible."
            )
        ).grid(row=1, column=5, columnspan=6, sticky="w", pady=(8, 0), padx=(8, 0))

        self.rules_frame = ttk.LabelFrame(main, text="4) Règles complémentaires", padding=10)
        self.rules_frame.grid(row=3, column=0, sticky="ew", pady=(0, 10))
        self.rules_frame.columnconfigure(3, weight=1)

        self.enable_email_validation_var = tk.BooleanVar(value=True)
        ttk.Checkbutton(
            self.rules_frame,
            text="Vérifier le format des emails",
            variable=self.enable_email_validation_var
        ).grid(row=0, column=0, sticky="w", padx=(0, 16))

        ttk.Label(self.rules_frame, text="Regex email :").grid(row=0, column=1, sticky="w")
        self.email_regex_var = tk.StringVar(value=EMAIL_REGEX_DEFAULT)
        ttk.Entry(self.rules_frame, textvariable=self.email_regex_var).grid(
            row=0, column=2, columnspan=2, sticky="ew", padx=(8, 0)
        )

        self.enable_length_validation_var = tk.BooleanVar(value=True)
        ttk.Checkbutton(
            self.rules_frame,
            text="Vérifier la longueur max depuis le header",
            variable=self.enable_length_validation_var
        ).grid(row=1, column=0, sticky="w", padx=(0, 16), pady=(8, 0))

        self.enable_type_validation_var = tk.BooleanVar(value=True)
        ttk.Checkbutton(
            self.rules_frame,
            text="Vérifier le type depuis le header",
            variable=self.enable_type_validation_var
        ).grid(row=2, column=0, sticky="w", padx=(0, 16), pady=(8, 0))

        ttk.Label(
            self.rules_frame,
            text=(
                "Types supportés : Texte/Text, Email, Numérique/Numeric/Number, "
                "Entier/Integer, Décimal/Decimal, Date, Booléen/Boolean, "
                "Téléphone/Phone, Alphanumérique/Alphanumeric"
            )
        ).grid(row=2, column=1, columnspan=3, sticky="w", pady=(8, 0))

        center = ttk.Frame(main)
        center.grid(row=4, column=0, sticky="nsew", pady=(0, 10))
        center.rowconfigure(0, weight=1)
        center.columnconfigure(0, weight=1)
        center.columnconfigure(1, weight=2)

        sheet_frame = ttk.LabelFrame(center, text="3) Onglets à vérifier", padding=10)
        sheet_frame.grid(row=0, column=0, sticky="nsew", padx=(0, 8))
        sheet_frame.rowconfigure(2, weight=1)
        sheet_frame.columnconfigure(0, weight=1)

        sheet_btn_row_1 = ttk.Frame(sheet_frame)
        sheet_btn_row_1.grid(row=0, column=0, sticky="ew", pady=(0, 6))

        ttk.Button(sheet_btn_row_1, text="Tout sélectionner", command=self.select_all_sheets).pack(
            side="left", padx=(0, 6)
        )
        ttk.Button(sheet_btn_row_1, text="Tout désélectionner", command=self.clear_sheet_selection).pack(
            side="left", padx=(0, 6)
        )
        ttk.Button(sheet_btn_row_1, text="Actualiser colonnes", command=self.refresh_detected_columns).pack(
            side="left"
        )

        self.sheet_info_var = tk.StringVar(value="Aucun onglet chargé.")
        ttk.Label(sheet_frame, textvariable=self.sheet_info_var).grid(row=1, column=0, sticky="w", pady=(0, 8))

        list_container = ttk.Frame(sheet_frame)
        list_container.grid(row=2, column=0, sticky="nsew")
        list_container.rowconfigure(0, weight=1)
        list_container.columnconfigure(0, weight=1)

        self.sheet_listbox = tk.Listbox(
            list_container,
            selectmode=tk.MULTIPLE,
            exportselection=False
        )
        self.sheet_listbox.grid(row=0, column=0, sticky="nsew")

        sheet_scrollbar = ttk.Scrollbar(list_container, orient="vertical", command=self.sheet_listbox.yview)
        sheet_scrollbar.grid(row=0, column=1, sticky="ns")
        self.sheet_listbox.config(yscrollcommand=sheet_scrollbar.set)
        self.sheet_listbox.bind("<<ListboxSelect>>", self.on_sheet_selection_changed)

        columns_frame = ttk.LabelFrame(center, text="4) Colonnes détectées / sélectionnées", padding=10)
        columns_frame.grid(row=0, column=1, sticky="nsew")
        columns_frame.rowconfigure(4, weight=1)
        columns_frame.columnconfigure(0, weight=1)

        top_row = ttk.Frame(columns_frame)
        top_row.grid(row=0, column=0, sticky="ew", pady=(0, 6))
        top_row.columnconfigure(1, weight=1)

        ttk.Label(top_row, text="Onglet à configurer :").grid(row=0, column=0, sticky="w", padx=(0, 8))

        self.sheet_to_configure_var = tk.StringVar()
        self.sheet_to_configure_combo = ttk.Combobox(
            top_row,
            textvariable=self.sheet_to_configure_var,
            state="readonly",
            values=[]
        )
        self.sheet_to_configure_combo.grid(row=0, column=1, sticky="ew")
        self.sheet_to_configure_combo.bind("<<ComboboxSelected>>", self.on_sheet_to_configure_changed)

        self.current_sheet_label_var = tk.StringVar(value="Aucun onglet sélectionné pour la configuration.")
        ttk.Label(columns_frame, textvariable=self.current_sheet_label_var).grid(
            row=1, column=0, sticky="w", pady=(0, 6)
        )

        legend_frame = ttk.Frame(columns_frame)
        legend_frame.grid(row=2, column=0, sticky="ew", pady=(0, 8))

        tk.Label(
            legend_frame,
            text="OBLIGATOIRE",
            bg="#FDE2E1",
            fg="#9C1C1C",
            padx=8,
            pady=3
        ).pack(side="left", padx=(0, 8))

        tk.Label(
            legend_frame,
            text="RECOMMANDÉE",
            bg="#FFF4CC",
            fg="#8A6A00",
            padx=8,
            pady=3
        ).pack(side="left", padx=(0, 8))

        columns_btn_row_1 = ttk.Frame(columns_frame)
        columns_btn_row_1.grid(row=3, column=0, sticky="ew", pady=(0, 8))

        ttk.Button(columns_btn_row_1, text="Tout cocher (onglet affiché)", command=self.select_all_columns_current_sheet).pack(
            side="left", padx=(0, 6)
        )
        ttk.Button(columns_btn_row_1, text="Tout décocher (onglet affiché)", command=self.clear_columns_current_sheet).pack(
            side="left", padx=(0, 6)
        )
        ttk.Button(columns_btn_row_1, text="Tout cocher (tous les onglets)", command=self.select_all_columns_all_sheets).pack(
            side="left", padx=(0, 6)
        )
        ttk.Button(columns_btn_row_1, text="Tout décocher (tous les onglets)", command=self.clear_columns_all_sheets).pack(
            side="left"
        )

        canvas_container = ttk.Frame(columns_frame)
        canvas_container.grid(row=4, column=0, sticky="nsew")
        canvas_container.rowconfigure(0, weight=1)
        canvas_container.columnconfigure(0, weight=1)

        self.columns_canvas = tk.Canvas(
            canvas_container,
            highlightthickness=0,
            bd=0,
            bg=bg_color
        )
        self.columns_canvas.grid(row=0, column=0, sticky="nsew")

        self.columns_scrollbar = ttk.Scrollbar(canvas_container, orient="vertical", command=self.columns_canvas.yview)
        self.columns_scrollbar.grid(row=0, column=1, sticky="ns")

        self.columns_canvas.configure(yscrollcommand=self.columns_scrollbar.set)

        self.columns_inner = ttk.Frame(self.columns_canvas)
        self.columns_window = self.columns_canvas.create_window((0, 0), window=self.columns_inner, anchor="nw")

        self.columns_inner.bind("<Configure>", self._on_columns_frame_configure)
        self.columns_canvas.bind("<Configure>", self._on_columns_canvas_configure)

        self.columns_info_var = tk.StringVar(value="Aucune colonne détectée.")
        ttk.Label(columns_frame, textvariable=self.columns_info_var).grid(row=5, column=0, sticky="w", pady=(8, 0))

        # info_frame = ttk.LabelFrame(main, text="7) Fonctionnement du contrôle", padding=10)
        # info_frame.grid(row=5, column=0, sticky="ew", pady=(0, 10))

        # info_text = (
        #     "- Si un fichier d'explication est chargé, la colonne E (Importance) est utilisée.\n"
        #     "- 'obligatoire' -> colonne classée en OBLIGATOIRE.\n"
        #     "- 'recommandé' -> colonne classée en RECOMMANDÉE.\n"
        #     "- Les colonnes obligatoires et recommandées sont cochées automatiquement.\n"
        #     "- Les autres colonnes restent disponibles mais décochées par défaut.\n"
        #     "- Les champs obligatoires vides = anomalie obligatoire.\n"
        #     "- Les champs recommandés vides = anomalie recommandée.\n"
        #     "- Les validations email / type / longueur restent actives."
        # )
        # ttk.Label(info_frame, text=info_text, justify="left").pack(anchor="w")

        action_frame = ttk.Frame(main)
        action_frame.grid(row=6, column=0, sticky="ew")
        action_frame.columnconfigure(0, weight=1)

        self.status_var = tk.StringVar(value="Prêt.")
        ttk.Label(action_frame, textvariable=self.status_var).grid(row=0, column=0, sticky="w")

        self.run_button = ttk.Button(
            action_frame,
            text="Lancer le contrôle et générer le reporting",
            command=self.run_check
        )
        self.run_button.grid(row=0, column=1, sticky="e")
        
        self.toggle_advanced_sections(False)

    def _on_columns_frame_configure(self, event):
        self.columns_canvas.configure(scrollregion=self.columns_canvas.bbox("all"))

    def _on_columns_canvas_configure(self, event):
        self.columns_canvas.itemconfigure(self.columns_window, width=event.width)

    def _on_main_frame_configure(self, event):
        self.main_canvas.configure(scrollregion=self.main_canvas.bbox("all"))

    def _on_main_canvas_configure(self, event):
        canvas_width = event.width
        canvas_height = event.height

        # Le contenu prend toute la largeur
        self.main_canvas.itemconfigure(self.main_window, width=canvas_width)

        # Le contenu prend au minimum toute la hauteur visible
        # Donc plus de grand blanc en bas si le contenu est plus petit que la fenêtre
        requested_height = self.content_frame.winfo_reqheight()
        final_height = max(canvas_height, requested_height)

        self.main_canvas.itemconfigure(self.main_window, height=final_height)
        self.main_canvas.configure(scrollregion=self.main_canvas.bbox("all"))

    def choose_file(self):
        path = filedialog.askopenfilename(
            title="Choisir un fichier Excel",
            filetypes=[
                ("Fichiers Excel", "*.xlsx *.xlsm"),
                ("Tous les fichiers", "*.*")
            ]
        )
        if not path:
            return

        self.file_path = path
        self.file_var.set(path)
        self.load_sheet_names()
        self.refresh_detected_columns()

    def choose_explanation_file(self):
        path = filedialog.askopenfilename(
            title="Choisir le fichier d'explication client",
            filetypes=[
                ("Fichiers Excel", "*.xlsx *.xlsm"),
                ("Tous les fichiers", "*.*")
            ]
        )
        if not path:
            return

        self.show_loading_popup("Chargement du fichier d'explication...")
        self.status_var.set("Chargement du fichier d'explication...")
        self.set_ui_enabled(False)

        thread = threading.Thread(
            target=self._load_explanation_file_worker,
            args=(path,),
            daemon=True
        )
        thread.start()

    def load_sheet_names(self):
        try:
            wb = load_workbook(self.file_path, read_only=True, data_only=False)
            self.sheet_names = wb.sheetnames
            wb.close()

            self.sheet_listbox.delete(0, tk.END)
            for sheet_name in self.sheet_names:
                self.sheet_listbox.insert(tk.END, sheet_name)

            self.select_all_sheets()
            self.sheet_info_var.set(f"{len(self.sheet_names)} onglet(s) chargé(s).")
            self.status_var.set("Fichier chargé.")
        except Exception as e:
            messagebox.showerror("Erreur", f"Impossible de lire le fichier Excel.\n\n{e}")

    def select_all_sheets(self):
        self.sheet_listbox.select_set(0, tk.END)
        self.on_sheet_selection_changed()

    def clear_sheet_selection(self):
        self.sheet_listbox.select_clear(0, tk.END)
        self.current_sheet_for_columns = None
        self.update_sheet_to_configure_combo()
        self.render_columns_for_current_sheet()
        self.on_sheet_selection_changed()

    def get_selected_sheets(self):
        indices = self.sheet_listbox.curselection()
        return [self.sheet_listbox.get(i) for i in indices]

    def on_sheet_selection_changed(self, event=None):
        selected_sheets = self.get_selected_sheets()
        self.sheet_info_var.set(f"{len(selected_sheets)} onglet(s) sélectionné(s).")
        self.update_sheet_to_configure_combo()

    def update_sheet_to_configure_combo(self):
        selected_sheets = self.get_selected_sheets()
        self.sheet_to_configure_combo["values"] = selected_sheets

        if not selected_sheets:
            self.sheet_to_configure_var.set("")
            self.current_sheet_for_columns = None
            self.render_columns_for_current_sheet()
            return

        if self.current_sheet_for_columns not in selected_sheets:
            self.current_sheet_for_columns = selected_sheets[0]

        self.sheet_to_configure_var.set(self.current_sheet_for_columns)
        self.render_columns_for_current_sheet()

    def on_sheet_to_configure_changed(self, event=None):
        selected_sheet = self.sheet_to_configure_var.get().strip()
        self.current_sheet_for_columns = selected_sheet if selected_sheet else None
        self.render_columns_for_current_sheet()

    def refresh_detected_columns(self):
        if not self.file_path:
            return

        try:
            header_row = int(self.header_row_var.get())
            if header_row < 1:
                raise ValueError
        except ValueError:
            messagebox.showwarning("Paramètre invalide", "La ligne des entêtes doit être un entier >= 1.")
            return

        detect_by_color = self.detect_by_color_var.get()
        detect_by_star = self.detect_by_star_var.get()

        target_rgb = None
        if self.use_custom_color_var.get():
            target_rgb = hex_to_rgb(self.custom_color_var.get())
            if not target_rgb:
                messagebox.showwarning(
                    "Couleur invalide",
                    "Veuillez entrer une couleur hex valide sur 6 caractères, par exemple FFFF00."
                )
                return

        use_explanation = self.use_explanation_file_var.get() and bool(
            self.explanation_data["mandatory"] or self.explanation_data["recommended"]
        )

        try:
            mandatory_marker_row = header_row + 1

            wb = load_workbook(self.file_path, data_only=False)

            new_detected = {}
            new_vars_by_sheet = {}
            new_categories_by_sheet = {}

            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]

                if use_explanation:
                    base_columns = get_all_non_empty_columns(ws, header_row)
                else:
                    base_columns = find_target_columns(
                        ws=ws,
                        header_row=header_row,
                        mandatory_marker_row=mandatory_marker_row,
                        detect_by_color=detect_by_color,
                        detect_by_star=detect_by_star,
                        target_rgb=target_rgb
                    )

                classified_columns = []
                categories_map = {}

                previous_vars = self.column_vars_by_sheet.get(sheet_name, {})
                current_vars = {}

                for col_idx, raw_header in base_columns:
                    category = classify_column_with_explanation(raw_header, self.explanation_data) if use_explanation else "mandatory"

                    if use_explanation and category == "other":
                        continue

                    default_value = True

                    previous_var = previous_vars.get(col_idx)
                    if previous_var is not None:
                        current_vars[col_idx] = tk.BooleanVar(value=previous_var.get())
                    else:
                        current_vars[col_idx] = tk.BooleanVar(value=default_value)

                    classified_columns.append({
                        "col_idx": col_idx,
                        "raw_header": raw_header,
                        "category": category
                    })
                    categories_map[col_idx] = category

                new_detected[sheet_name] = classified_columns
                new_vars_by_sheet[sheet_name] = current_vars
                new_categories_by_sheet[sheet_name] = categories_map

            wb.close()

            self.detected_columns_by_sheet = new_detected
            self.column_vars_by_sheet = new_vars_by_sheet
            self.column_categories_by_sheet = new_categories_by_sheet

            self.update_sheet_to_configure_combo()

            if use_explanation:
                self.status_var.set("Colonnes classées via le fichier d'explication (obligatoire / recommandée / autre).")
            else:
                mode_label = build_detection_mode_label(detect_by_color, detect_by_star, target_rgb)
                self.status_var.set(f"Colonnes détectées ({mode_label}).")

        except Exception as e:
            messagebox.showerror("Erreur", f"Impossible de détecter les colonnes.\n\n{e}")

    def render_columns_for_current_sheet(self):
        for widget in self.columns_inner.winfo_children():
            widget.destroy()

        sheet_name = self.current_sheet_for_columns

        if not sheet_name:
            self.current_sheet_label_var.set("Aucun onglet sélectionné pour la configuration.")
            self.columns_info_var.set("Aucun onglet sélectionné.")
            return

        self.current_sheet_label_var.set(f"Onglet configuré : {sheet_name}")

        detected_columns = self.detected_columns_by_sheet.get(sheet_name, [])
        var_map = self.column_vars_by_sheet.get(sheet_name, {})

        if not detected_columns:
            ttk.Label(
                self.columns_inner,
                text="Aucune colonne détectée sur cet onglet."
            ).pack(anchor="w", padx=4, pady=4)
            self.columns_info_var.set(f"{sheet_name} : 0 colonne détectée.")
            return

        mandatory_cols = [c for c in detected_columns if c["category"] == "mandatory"]
        recommended_cols = [c for c in detected_columns if c["category"] == "recommended"]

        self._render_column_section(
            title="Colonnes OBLIGATOIRES",
            bg="#FDE2E1",
            fg="#9C1C1C",
            columns=mandatory_cols,
            var_map=var_map
        )

        self._render_column_section(
            title="Colonnes RECOMMANDÉES",
            bg="#FFF4CC",
            fg="#8A6A00",
            columns=recommended_cols,
            var_map=var_map
        )

        self.update_columns_info()

    def _render_column_section(self, title, bg, fg, columns, var_map):
        if not columns:
            return

        section_frame = ttk.Frame(self.columns_inner)
        section_frame.pack(fill="x", padx=4, pady=(4, 8), anchor="w")

        tk.Label(
            section_frame,
            text=title,
            bg=bg,
            fg=fg,
            anchor="w",
            font=("Segoe UI", 10, "bold"),
            padx=8,
            pady=5
        ).pack(fill="x", anchor="w")

        inner = ttk.Frame(section_frame)
        inner.pack(fill="x", padx=(6, 0), pady=(4, 0))

        for item in columns:
            col_idx = item["col_idx"]
            raw_header = item["raw_header"]

            var = var_map.get(col_idx)
            if var is None:
                var = tk.BooleanVar(value=True)
                var_map[col_idx] = var

            meta = parse_header_metadata(raw_header)
            label = meta["field_name"] or clean_header_name(raw_header)

            extra = []
            if meta["field_type"]:
                extra.append(f"Type: {meta['field_type']}")
            if meta["max_length"] is not None:
                extra.append(f"Longueur: {meta['max_length']}")
            if is_email_header(raw_header):
                extra.append("Email")

            suffix = f" [{' | '.join(extra)}]" if extra else ""
            text = f"Colonne {col_idx} — {label}{suffix}"

            cb = ttk.Checkbutton(
                inner,
                text=text,
                variable=var,
                command=self.update_columns_info
            )
            cb.pack(anchor="w", pady=2)

    def update_columns_info(self):
        sheet_name = self.current_sheet_for_columns
        if not sheet_name:
            self.columns_info_var.set("Aucun onglet sélectionné.")
            return

        detected_columns = self.detected_columns_by_sheet.get(sheet_name, [])
        var_map = self.column_vars_by_sheet.get(sheet_name, {})

        total = len(detected_columns)
        selected = 0
        mandatory_selected = 0
        recommended_selected = 0
        other_selected = 0

        for item in detected_columns:
            col_idx = item["col_idx"]
            category = item["category"]

            var = var_map.get(col_idx)
            if var and var.get():
                selected += 1
                if category == "mandatory":
                    mandatory_selected += 1
                elif category == "recommended":
                    recommended_selected += 1
                else:
                    other_selected += 1

        self.columns_info_var.set(
            f"{sheet_name} : {total} colonne(s), "
            f"{selected} cochée(s) "
            f"(Obligatoires: {mandatory_selected} | Recommandées: {recommended_selected} | Autres: {other_selected})."
        )

    def select_all_columns_current_sheet(self):
        sheet_name = self.current_sheet_for_columns
        if not sheet_name:
            return

        for var in self.column_vars_by_sheet.get(sheet_name, {}).values():
            var.set(True)

        self.render_columns_for_current_sheet()

    def clear_columns_current_sheet(self):
        sheet_name = self.current_sheet_for_columns
        if not sheet_name:
            return

        for var in self.column_vars_by_sheet.get(sheet_name, {}).values():
            var.set(False)

        self.render_columns_for_current_sheet()

    def select_all_columns_all_sheets(self):
        for var_map in self.column_vars_by_sheet.values():
            for var in var_map.values():
                var.set(True)

        self.render_columns_for_current_sheet()

    def clear_columns_all_sheets(self):
        for var_map in self.column_vars_by_sheet.values():
            for var in var_map.values():
                var.set(False)

        self.render_columns_for_current_sheet()

    def get_selected_columns_by_sheet(self, selected_sheets):
        result = {}

        for sheet_name in selected_sheets:
            detected_columns = self.detected_columns_by_sheet.get(sheet_name, [])
            var_map = self.column_vars_by_sheet.get(sheet_name, {})

            selected_columns = []
            for item in detected_columns:
                col_idx = item["col_idx"]
                raw_header = item["raw_header"]
                category = item["category"]

                var = var_map.get(col_idx)
                if var and var.get():
                    selected_columns.append({
                        "col_idx": col_idx,
                        "raw_header": raw_header,
                        "category": category
                    })

            result[sheet_name] = selected_columns

        return result

    def run_check(self):
        if not self.file_path:
            messagebox.showwarning("Fichier manquant", "Veuillez choisir un fichier Excel.")
            return

        selected_sheets = self.get_selected_sheets()
        if not selected_sheets:
            messagebox.showwarning("Aucun onglet", "Veuillez sélectionner au moins un onglet.")
            return

        try:
            header_row = int(self.header_row_var.get())
            if header_row < 1:
                raise ValueError
        except ValueError:
            messagebox.showwarning("Paramètre invalide", "La ligne des entêtes doit être un entier >= 1.")
            return

        email_regex = self.email_regex_var.get().strip()
        if self.enable_email_validation_var.get():
            try:
                re.compile(email_regex)
            except re.error as e:
                messagebox.showwarning(
                    "Regex invalide",
                    f"La regex email est invalide.\n\n{e}"
                )
                return

        if not self.detected_columns_by_sheet:
            self.refresh_detected_columns()

        selected_columns_by_sheet = self.get_selected_columns_by_sheet(selected_sheets)
        total_selected_columns = sum(len(cols) for cols in selected_columns_by_sheet.values())

        if total_selected_columns == 0:
            messagebox.showwarning(
                "Aucune colonne sélectionnée",
                "Aucune colonne n'est cochée sur les onglets sélectionnés."
            )
            return

        self.show_loading_popup("Génération du reporting...")
        self.status_var.set("Génération du reporting en cours...")
        self.set_ui_enabled(False)

        thread = threading.Thread(
            target=self._generate_report_worker,
            kwargs={
                "selected_sheets": selected_sheets,
                "header_row": header_row,
                "selected_columns_by_sheet": selected_columns_by_sheet,
                "email_regex": email_regex,
                "enable_email_validation": self.enable_email_validation_var.get(),
                "enable_length_validation": self.enable_length_validation_var.get(),
                "enable_type_validation": self.enable_type_validation_var.get(),
            },
            daemon=True
        )
        thread.start()

    def _generate_report_worker(
        self,
        selected_sheets,
        header_row,
        selected_columns_by_sheet,
        email_regex,
        enable_email_validation,
        enable_length_validation,
        enable_type_validation
    ):
        try:
            output_path = generate_report(
                input_file=self.file_path,
                selected_sheets=selected_sheets,
                header_row=header_row,
                selected_columns_by_sheet=selected_columns_by_sheet,
                enable_email_validation=enable_email_validation,
                email_regex=email_regex,
                enable_length_validation=enable_length_validation,
                enable_type_validation=enable_type_validation
            )

            self.root.after(0, lambda path=output_path: self._on_report_generated(path))

        except Exception as e:
            self.root.after(0, lambda error=e: self._on_report_error(error))


    def _on_report_generated(self, output_path):
        self.hide_loading_popup()
        self.set_ui_enabled(True)

        self.status_var.set("Reporting généré.")

        messagebox.showinfo(
            "Succès",
            f"Reporting généré avec succès :\n\n{output_path}"
        )


    def _on_report_error(self, error):
        self.hide_loading_popup()
        self.set_ui_enabled(True)

        self.status_var.set("Erreur pendant la génération du reporting.")

        messagebox.showerror(
            "Erreur",
            f"Une erreur est survenue pendant le contrôle.\n\n{error}"
        )

    def toggle_advanced_sections(self, show: bool):
        if show:
            self.param_frame.grid()
            self.rules_frame.grid()
        else:
            self.param_frame.grid_remove()
            self.rules_frame.grid_remove()
    
    def show_loading_popup(self, message="Chargement en cours..."):
        if hasattr(self, "loading_popup") and self.loading_popup and self.loading_popup.winfo_exists():
            return

        self.loading_popup = tk.Toplevel(self.root)
        self.loading_popup.title("Veuillez patienter")
        try:
            self.loading_popup.iconbitmap(self.resource_path("app2.ico"))
        except Exception:
            pass
        self.loading_popup.transient(self.root)
        self.loading_popup.grab_set()
        self.loading_popup.resizable(False, False)

        # Empêche la fermeture manuelle pendant le traitement
        self.loading_popup.protocol("WM_DELETE_WINDOW", lambda: None)

        frame = ttk.Frame(self.loading_popup, padding=20)
        frame.pack(fill="both", expand=True)

        ttk.Label(frame, text=message, font=("Segoe UI", 11, "bold")).pack(pady=(0, 10))
        ttk.Label(frame, text="L'application travaille, merci de patienter.").pack()

        progress = ttk.Progressbar(frame, mode="indeterminate", length=250)
        progress.pack(pady=(12, 0))
        progress.start(10)

        self.loading_progress = progress

        self.loading_popup.update_idletasks()

        # Centrage
        popup_width = self.loading_popup.winfo_reqwidth()
        popup_height = self.loading_popup.winfo_reqheight()
        root_x = self.root.winfo_rootx()
        root_y = self.root.winfo_rooty()
        root_w = self.root.winfo_width()
        root_h = self.root.winfo_height()

        x = root_x + (root_w // 2) - (popup_width // 2)
        y = root_y + (root_h // 2) - (popup_height // 2)

        self.loading_popup.geometry(f"+{x}+{y}")

    def hide_loading_popup(self):
        if hasattr(self, "loading_progress") and self.loading_progress:
            try:
                self.loading_progress.stop()
            except Exception:
                pass

        if hasattr(self, "loading_popup") and self.loading_popup:
            try:
                self.loading_popup.grab_release()
            except Exception:
                pass
            try:
                self.loading_popup.destroy()
            except Exception:
                pass

        self.loading_popup = None
        self.loading_progress = None

    def set_ui_enabled(self, enabled: bool):
        state = "normal" if enabled else "disabled"

        widgets = [
            getattr(self, "sheet_listbox", None),
            getattr(self, "sheet_to_configure_combo", None),
            getattr(self, "run_button", None),
        ]

        for widget in widgets:
            if widget is not None:
                try:
                    widget.configure(state=state)
                except Exception:
                    pass

    def _load_explanation_file_worker(self, path):
        try:
            explanation_data = parse_explanation_workbook(path)
            mandatory_count = len(explanation_data["mandatory"])
            recommended_count = len(explanation_data["recommended"])

            self.root.after(0, lambda: self._on_explanation_file_loaded(
                path,
                explanation_data,
                mandatory_count,
                recommended_count
            ))
        except Exception as e:
            self.root.after(0, lambda error=e: self._on_explanation_file_error(error))

    def _on_explanation_file_loaded(self, path, explanation_data, mandatory_count, recommended_count):
        self.explanation_file_path = path
        self.explanation_file_var.set(path)
        self.explanation_data = explanation_data

        self.explanation_info_var.set(
            f"Fichier d'explication chargé : {mandatory_count} champ(s) obligatoire(s), "
            f"{recommended_count} champ(s) recommandé(s)."
        )

        self.status_var.set("Fichier d'explication chargé.")
        self.hide_loading_popup()
        self.set_ui_enabled(True)

        self.refresh_detected_columns()

    def _on_explanation_file_error(self, error):
        self.hide_loading_popup()
        self.set_ui_enabled(True)
        messagebox.showerror(
            "Erreur",
            f"Impossible de lire le fichier d'explication.\n\n{error}"
        )