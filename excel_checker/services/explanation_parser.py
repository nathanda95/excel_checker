from openpyxl import load_workbook

from excel_checker.utils.text_utils import (
    normalize_key,
    normalize_text_for_compare,
    parse_header_metadata,
    clean_header_name,
    parse_importance_label,
)


def parse_explanation_workbook(path: str):
    wb = load_workbook(path, data_only=True, read_only=True)

    if not wb.sheetnames:
        wb.close()
        return {
            "mandatory": set(),
            "recommended": set(),
            "by_group": {}
        }

    ws = wb[wb.sheetnames[0]]

    mandatory = set()
    recommended = set()
    by_group = {}

    current_group = ""

    for row_idx in range(1, ws.max_row + 1):
        group_value = ws.cell(row=row_idx, column=2).value
        description_value = ws.cell(row=row_idx, column=4).value
        importance_value = ws.cell(row=row_idx, column=5).value

        group_text = normalize_text_for_compare(group_value)
        description_text = normalize_text_for_compare(description_value)
        importance_type = parse_importance_label(importance_value)

        if group_text and not description_text:
            current_group = group_text
            if current_group not in by_group:
                by_group[current_group] = {
                    "mandatory": set(),
                    "recommended": set()
                }

        if not description_text:
            continue

        normalized_desc = normalize_key(description_text)
        if not normalized_desc:
            continue

        if importance_type == "mandatory":
            mandatory.add(normalized_desc)
            if current_group:
                by_group.setdefault(current_group, {"mandatory": set(), "recommended": set()})
                by_group[current_group]["mandatory"].add(normalized_desc)

        elif importance_type == "recommended":
            recommended.add(normalized_desc)
            if current_group:
                by_group.setdefault(current_group, {"mandatory": set(), "recommended": set()})
                by_group[current_group]["recommended"].add(normalized_desc)

    wb.close()

    recommended = recommended - mandatory

    for group_name, group_data in by_group.items():
        group_data["recommended"] = group_data["recommended"] - group_data["mandatory"]

    return {
        "mandatory": mandatory,
        "recommended": recommended,
        "by_group": by_group
    }


def classify_column_with_explanation(raw_header, explanation_data):
    meta = parse_header_metadata(raw_header)
    field_name = meta["field_name"] or clean_header_name(raw_header)
    key = normalize_key(field_name)

    if not key:
        return "other"

    if key in explanation_data["mandatory"]:
        return "mandatory"

    if key in explanation_data["recommended"]:
        return "recommended"

    return "other"
