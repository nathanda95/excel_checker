import re
from pathlib import Path

from openpyxl import load_workbook, Workbook  # pyright: ignore[reportMissingModuleSource]
from openpyxl.styles import Font, PatternFill, Alignment  # pyright: ignore[reportMissingModuleSource]

from excel_checker.config import EMAIL_REGEX_DEFAULT
from excel_checker.utils.text_utils import (
    is_empty,
    normalize_sheet_title,
    normalize_text_for_compare,
    parse_header_metadata,
    clean_header_name,
)
from excel_checker.utils.validation import (
    is_email_header,
    is_valid_email,
    is_type_valid,
)
from excel_checker.services.postal_code_service import (
    load_postal_code_rules,
    is_postal_code_header,
    is_country_header,
    validate_postal_code_for_country,
)
from excel_checker.services.phone_number_service import (
    load_phone_number_rules,
    is_phone_header,
    validate_phone_for_country,
)

ERROR_CATEGORY_META = {
    "mandatory_missing": {
        "label": "Donnée obligatoire manquante",
        "priority": 1,
        "fill": "FDE2E1",
        "font": "9C1C1C",
    },
    "recommended_missing": {
        "label": "Donnée recommandée manquante",
        "priority": 2,
        "fill": "FFF4CC",
        "font": "8A6A00",
    },
    "invalid_data": {
        "label": "Donnée incorrecte",
        "priority": 3,
        "fill": "DDEBF7",
        "font": "1F4E78",
    },
}


def row_has_any_data(ws, row_number: int) -> bool:
    for cell in ws[row_number]:
        if not is_empty(cell.value):
            return True
    return False


def make_anomaly(category_code: str, message: str) -> dict:
    meta = ERROR_CATEGORY_META[category_code]
    return {
        "category_code": category_code,
        "category_label": meta["label"],
        "priority": meta["priority"],
        "message": message,
    }


def sort_anomalies(anomalies: list[dict]) -> list[dict]:
    return sorted(anomalies, key=lambda item: (item["priority"], item["message"].lower()))


def get_main_category(anomalies: list[dict]) -> tuple[str, str, int]:
    if not anomalies:
        return "", "", 999

    first = sort_anomalies(anomalies)[0]
    return first["category_code"], first["category_label"], first["priority"]


def build_errors_for_sheet(
    ws,
    selected_columns,
    data_start_row: int,
    enable_email_validation: bool = True,
    email_regex: str = EMAIL_REGEX_DEFAULT,
    enable_length_validation: bool = True,
    enable_type_validation: bool = True
):
    if not selected_columns:
        return []

    errors = []
    compiled_email_regex = re.compile(email_regex)
    postal_rules = load_postal_code_rules()
    phone_rules = load_phone_number_rules()

    country_columns = []
    postal_columns = []
    phone_columns = []

    # Colonnes pays détectées dans toute la feuille
    # pour ne pas dépendre du fait qu'elles soient cochées ou non
    header_row_num = data_start_row - 1

    for col_idx in range(1, ws.max_column + 1):
        raw_header = ws.cell(row=header_row_num, column=col_idx).value

        if is_country_header(raw_header):
            country_columns.append({
                "col_idx": col_idx,
                "raw_header": raw_header,
            })

    # Colonnes postal / téléphone détectées uniquement parmi les colonnes cochées
    for item in selected_columns:
        raw_header = item["raw_header"]

        if is_postal_code_header(raw_header):
            postal_columns.append(item)

        if is_phone_header(raw_header):
            phone_columns.append(item)

    def get_country_value_for_row(row_num: int, target_col_idx: int):
        """
        Cherche la colonne pays la plus pertinente pour la colonne cible.
        Priorité :
        1) colonne pays à gauche la plus proche ET non vide sur la ligne
        2) sinon colonne pays la plus proche tout court ET non vide sur la ligne
        3) sinon None
        """
        if not country_columns:
            return None, None

        # 1) candidates à gauche, triées de la plus proche à la plus éloignée
        left_candidates = sorted(
            [c for c in country_columns if c["col_idx"] < target_col_idx],
            key=lambda c: c["col_idx"],
            reverse=True
        )

        for candidate in left_candidates:
            candidate_value = ws.cell(row=row_num, column=candidate["col_idx"]).value
            if not is_empty(candidate_value):
                return candidate_value, candidate["col_idx"]

        # 2) sinon on prend la plus proche tout court, mais uniquement si non vide
        all_candidates = sorted(
            country_columns,
            key=lambda c: abs(c["col_idx"] - target_col_idx)
        )

        for candidate in all_candidates:
            candidate_value = ws.cell(row=row_num, column=candidate["col_idx"]).value
            if not is_empty(candidate_value):
                return candidate_value, candidate["col_idx"]

        return None, None

    print("DEBUG phone_columns:", phone_columns)
    print("DEBUG postal_columns:", postal_columns)
    print("DEBUG country_columns:", country_columns)

    for row_num in range(data_start_row, ws.max_row + 1):
        if not row_has_any_data(ws, row_num):
            continue

        anomalies = []

        for item in selected_columns:
            col_idx = item["col_idx"]
            raw_header = item["raw_header"]
            category = item.get("category", "mandatory")

            meta = parse_header_metadata(raw_header)
            field_name = meta["field_name"] or clean_header_name(raw_header)
            field_type = meta["field_type"]
            max_length = meta["max_length"]

            value = ws.cell(row=row_num, column=col_idx).value

            print(
                "DEBUG",
                ws.title,
                "row=", row_num,
                "col=", col_idx,
                "field=", field_name,
                "value=", repr(value)
            )

            # 1) Champ vide
            if is_empty(value):
                if category == "mandatory":
                    anomalies.append(make_anomaly(
                        "mandatory_missing",
                        f"Champ obligatoire manquant : {field_name}"
                    ))
                elif category == "recommended":
                    anomalies.append(make_anomaly(
                        "recommended_missing",
                        f"Champ recommandé manquant : {field_name}"
                    ))
                else:
                    anomalies.append(make_anomaly(
                        "invalid_data",
                        f"Champ sélectionné manquant : {field_name}"
                    ))
                continue

            text_value = normalize_text_for_compare(value)

            # 2) Longueur max dépassée
            if enable_length_validation and max_length is not None:
                if len(text_value) > max_length:
                    anomalies.append(make_anomaly(
                        "invalid_data",
                        f"Longueur dépassée : {field_name} ({len(text_value)} > {max_length})"
                    ))

            # 3) Validation email
            if enable_email_validation and is_email_header(raw_header):
                if not is_valid_email(text_value, compiled_email_regex):
                    anomalies.append(make_anomaly(
                        "invalid_data",
                        f"Email invalide : {field_name} ({text_value})"
                    ))

            # 4) Validation type
            if enable_type_validation and field_type:
                ok_type, type_label = is_type_valid(
                    field_type,
                    text_value,
                    email_regex=email_regex
                )
                if not ok_type:
                    anomalies.append(make_anomaly(
                        "invalid_data",
                        f"Type invalide : {field_name} (attendu: {type_label}, valeur: {text_value})"
                    ))

            # 5) Validation code postal selon le pays
            if postal_rules and is_postal_code_header(raw_header):
                country_value, best_country_col_idx = get_country_value_for_row(row_num, col_idx)

                ok_postal, matched_rule = validate_postal_code_for_country(
                    postal_code=text_value,
                    country_value=country_value,
                    postal_rules=postal_rules
                )

                print(
                    "POSTAL DEBUG |",
                    "sheet=", ws.title,
                    "| row=", row_num,
                    "| postal_header=", repr(raw_header),
                    "| postal_value=", repr(text_value),
                    "| country_col=", best_country_col_idx,
                    "| country_value=", repr(country_value),
                    "| matched_rule=", matched_rule,
                    "| ok_postal=", ok_postal
                )

                if matched_rule is not None and not ok_postal:
                    anomalies.append(make_anomaly(
                        "invalid_data",
                        (
                            f"Code postal invalide : {field_name} "
                            f"({text_value}) pour pays {matched_rule['label']}"
                            + (f" — ex: {matched_rule['example']}" if matched_rule.get("example") else "")
                        )
                    ))

            # 6) Validation téléphone selon le pays
            if phone_rules and is_phone_header(raw_header):
                country_value, best_country_col_idx = get_country_value_for_row(row_num, col_idx)

                ok_phone, matched_rule = validate_phone_for_country(
                    phone_value=text_value,
                    country_value=country_value,
                    phone_rules=phone_rules
                )

                print(
                    "PHONE DEBUG |",
                    "sheet=", ws.title,
                    "| row=", row_num,
                    "| phone_header=", repr(raw_header),
                    "| phone_value=", repr(text_value),
                    "| country_col=", best_country_col_idx,
                    "| country_value=", repr(country_value),
                    "| matched_rule=", matched_rule,
                    "| ok_phone=", ok_phone
                )

                if matched_rule is not None and not ok_phone:
                    anomalies.append(make_anomaly(
                        "invalid_data",
                        (
                            f"Téléphone invalide : {field_name} "
                            f"({text_value}) pour pays {matched_rule['label']}"
                            + (
                                f" — indicatif: {matched_rule['country_calling_code']}"
                                if matched_rule.get("country_calling_code")
                                else ""
                            )
                            + (f" — ex: {matched_rule['example']}" if matched_rule.get("example") else "")
                        )
                    ))

        if anomalies:
            anomalies = sort_anomalies(anomalies)
            main_code, main_label, main_priority = get_main_category(anomalies)

            errors.append({
                "sheet": ws.title,
                "row": row_num,
                "anomalies": anomalies,
                "main_category_code": main_code,
                "main_category_label": main_label,
                "main_priority": main_priority,
            })

    errors.sort(key=lambda err: (err["main_priority"], err["row"]))
    return errors


def apply_category_style(cell, category_code: str):
    meta = ERROR_CATEGORY_META.get(category_code)
    if not meta:
        return

    cell.fill = PatternFill(fill_type="solid", fgColor=meta["fill"])
    cell.font = Font(color=meta["font"])
    cell.alignment = Alignment(vertical="top", wrap_text=True)


def write_sheet_report(ws_out, errors, include_summary=True, include_detail_columns=True):
    max_missing = max((len(err["anomalies"]) for err in errors), default=0)

    headers = [
        "Onglet impacté",
        "Ligne Excel",
        "Nombre d'anomalies",
        "Type principal d'anomalie",
    ]

    if include_summary:
        headers.append("Résumé des anomalies")

    if include_detail_columns:
        headers += [f"Anomalie {i}" for i in range(1, max_missing + 1)]

    ws_out.append(headers)


    header_fill = PatternFill(fill_type="solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col_idx, _ in enumerate(headers, start=1):
        cell = ws_out.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    for err in errors:
        anomaly_messages = [a["message"] for a in err["anomalies"]]

        row = [
            err["sheet"],
            err["row"],
            len(err["anomalies"]),
            err["main_category_label"],
        ]

        if include_summary:
            summary = ", ".join(anomaly_messages)
            row.append(summary)

        if include_detail_columns:
            row.extend(anomaly_messages)

        ws_out.append(row)
        current_row = ws_out.max_row

        # Centrer la colonne "Nombre d'anomalies" (colonne 3)
        ws_out.cell(row=current_row, column=3).alignment = Alignment(
            horizontal="center",
            vertical="center"
        )

        # Colorer la catégorie principale
        apply_category_style(ws_out.cell(row=current_row, column=4), err["main_category_code"])

        # Colorer le résumé global selon la catégorie principale
        if include_summary:
            summary_col = 5
            apply_category_style(ws_out.cell(row=current_row, column=summary_col), err["main_category_code"])

        # Colorer chaque anomalie individuellement
        if include_detail_columns:
            detail_start_col = 5 if not include_summary else 6
            for idx, anomaly in enumerate(err["anomalies"]):
                apply_category_style(
                    ws_out.cell(row=current_row, column=detail_start_col + idx),
                    anomaly["category_code"]
                )

    for row in ws_out.iter_rows(min_row=2):
        for cell in row:
            if cell.alignment is None or not cell.alignment.wrap_text:
                cell.alignment = Alignment(vertical="top", wrap_text=True)

    base_widths = {
        1: 28,
        2: 12,
        3: 18,
        4: 28,
    }

    for col_idx, width in base_widths.items():
        ws_out.column_dimensions[ws_out.cell(row=1, column=col_idx).column_letter].width = width

    next_col = 5

    if include_summary:
        ws_out.column_dimensions[ws_out.cell(row=1, column=next_col).column_letter].width = 70
        next_col += 1

    if include_detail_columns:
        for col_idx in range(next_col, ws_out.max_column + 1):
            col_letter = ws_out.cell(row=1, column=col_idx).column_letter
            max_length = 0
            for row_idx in range(1, ws_out.max_row + 1):
                value = ws_out.cell(row=row_idx, column=col_idx).value
                value_str = "" if value is None else str(value)
                if len(value_str) > max_length:
                    max_length = len(value_str)
            ws_out.column_dimensions[col_letter].width = min(max(max_length + 2, 24), 48)

    ws_out.freeze_panes = "A2"
    ws_out.auto_filter.ref = ws_out.dimensions
    ws_out.row_dimensions[1].height = 30


def generate_report(
    input_file: str,
    selected_sheets: list[str],
    header_row: int,
    selected_columns_by_sheet: dict,
    enable_email_validation: bool = True,
    email_regex: str = EMAIL_REGEX_DEFAULT,
    enable_length_validation: bool = True,
    enable_type_validation: bool = True
):
    data_start_row = header_row + 1

    wb_in = load_workbook(input_file, data_only=True)
    wb_out = Workbook()

    default_sheet = wb_out.active
    wb_out.remove(default_sheet)

    all_errors = []

    for sheet_name in selected_sheets:
        if sheet_name not in wb_in.sheetnames:
            continue

        selected_columns = selected_columns_by_sheet.get(sheet_name, [])
        if not selected_columns:
            continue

        ws = wb_in[sheet_name]
        sheet_errors = build_errors_for_sheet(
            ws=ws,
            selected_columns=selected_columns,
            data_start_row=data_start_row,
            enable_email_validation=enable_email_validation,
            email_regex=email_regex,
            enable_length_validation=enable_length_validation,
            enable_type_validation=enable_type_validation
        )
        all_errors.extend(sheet_errors)

        if sheet_errors:
            out_sheet_name = normalize_sheet_title(sheet_name)
            ws_out = wb_out.create_sheet(title=out_sheet_name)
            write_sheet_report(
                ws_out,
                sheet_errors,
                include_summary=False,
                include_detail_columns=True
            )

    ws_global = wb_out.create_sheet(title="Reporting_Global", index=0)

    if all_errors:
        all_errors.sort(key=lambda err: (err["main_priority"], err["sheet"].lower(), err["row"]))
        write_sheet_report(
            ws_global,
            all_errors,
            include_summary=True,
            include_detail_columns=False
        )
    else:
        ws_global["A1"] = "Aucune erreur détectée sur les onglets/colonnes sélectionnés."
        ws_global.column_dimensions["A"].width = 80

    input_path = Path(input_file)
    output_path = input_path.with_name(f"{input_path.stem}_reporting.xlsx")
    wb_out.save(output_path)
    wb_in.close()

    return str(output_path)